import openpyxl
from openpyxl import load_workbook
import datetime

wb = load_workbook("test_1.xlsx")
sheet_1 = wb['active_tabs']
sheet_2 = wb['checkout']

active_input_list = []
column_data = []
active_tables_nums = []
orders = []
index_nums = []
log = []
n = 2


import datetime

def add_input():
    # input the table number
    table_num = input("Input the table number: ")
    table_num = table_num.zfill(2)
    active_input_list.append(table_num)
    timestamp = datetime.datetime.now()
    active_input_list.append(timestamp)
    # input the order as string then add it to the order list
    order = input("add the order: ")
    active_input_list.append(order)

    # check if there are other orders for the same table
    add_order = input("do u want to add another order? y/n: ")

    # while loop for adding other orders to the same table
    while add_order == "y":
        order = input("add the order: ")
        active_input_list.append(order)
        add_order = input("do u want to add another order? y/n: ")


# function that adds items to the Excel file
def active_tabs_input():
    add_input()
    sheet_1 = wb['active_tabs']
    # max rows to decide the starting row as max_row+1
    num_rows = sheet_1.max_row
    # i = index so the items from the list take the same index on the data sheet
    for i, item in enumerate(active_input_list, start=0):
        sheet_1.cell(row=num_rows + 1, column=i + 1).value = item
    wb.save("test_1.xlsx")
    # clearing the list so we can start a new cycle
    active_input_list.clear()


# function to find the active table numbers and add them to active_tables_list
def find_checkout_table():
    global active_tables_nums
    for row in sheet_1.iter_rows(min_col=1, max_col=1, min_row=2, values_only=True):
        for cell in row:
            active_tables_nums.append(cell)


# indexing the active_table_list to find the wanna checkout table row number in xlsx
def find_checkout_table_row_number(table_num, n):
    for num in active_tables_nums:
        if num == table_num:
            index_nums.append(n)
            n += 1
        else:
            n += 1

# copy the orders from xlsx to orders_list
def create_order_list_for_checkout():
    global orders
    for number in index_nums:
        for column in sheet_1.iter_rows(min_row=number, max_row=number, min_col=3, values_only=True):
            for order in column:
                orders.append(order)
                orders = [x for x in orders if x is not None]


def delete_checkout_row_from_active():
    print(index_nums)
    print(active_tables_nums)

    # Create a copy of index_nums
    index_nums_copy = list(index_nums)

    # Iterate over the copy and delete rows from sheet_1
    for w in index_nums_copy:
        sheet_1.delete_rows(w)

    index_nums.clear()
    print(index_nums)



# function to order list to the checkout sheet
def move_to_checkout_sheet():
    # i = index so the items from the list take the same index on the data sheet
    for i, order in enumerate(orders, start=0):
        sheet_2.cell(row=1, column=i + 1).value = order
    wb.save("test_1.xlsx")


# activating the checkout function
def checkout(n):
    global table_num
    table_num = input("input the table number: ")
    table_num = table_num.zfill(2)
    find_checkout_table()
    find_checkout_table_row_number(table_num, n)
    create_order_list_for_checkout()
    print(orders)
    move_to_checkout_sheet()
    delete_checkout_row_from_active()
    orders.clear()
    active_tables_nums.clear()


def create_table_log():
    global log
    for i in index_nums:
        for col in range(1, sheet_1.max_column + 1):
            element = sheet_1.cell(row=i, column=col).value
            log.append(element)
            log = [y for y in log if y is not None]
        print(log)
        active_tables_nums.clear()
        index_nums.clear()
        log.clear()


# function to show the orders from a specific table with timestamp without appending it anywhere
def check_table_log(n):
    table_num = input("input the table number: ")
    table_num = table_num.zfill(2)
    find_checkout_table()
    find_checkout_table_row_number(table_num, n)
    create_table_log()
    log.clear()




# choosing a task to start the program with
def choose_task(n):
    global orders
    global table_num
    choose_task = int(input("Choose a Task: type 1 or 2 or 3 \n"
                            "Register an order..........(1) \n"
                            "Checkout...................(2) \n"
                            "Check table log............(3): "))
    match choose_task:
        case 1:
            active_tabs_input()
        case 2:
            checkout(n)
        case 3:
            check_table_log(n)


# make the program work forever
while True:
    choose_task(n)
