import openpyxl

checkout_table_num= float(input("input the table number: "))

wb = openpyxl.load_workbook(filename = 'test_1.xlsx')
sheet_1 = wb['active_tabs']
sheet_2 = wb['checkout']

column_data = []
active_tables_nums =[]
orders=[]
index_nums = []
n=2

def find_checkout_table():
    for row in sheet_1.iter_rows(min_col=1, max_col=1, min_row=2, values_only=True):
        for cell in row:
            active_tables_nums.append(cell)

def find_checkout_table_row_number(n):
    for num in active_tables_nums:
        if num == checkout_table_num:
            index_nums.append(n)
            n +=1
        else: n +=1

def create_order_list_for_checkout():
    for number in index_nums:
        for column in sheet_1.iter_rows(min_row=number, max_row=number, min_col=2, values_only=True):
            for order in column:
                orders.append(order)


find_checkout_table()
find_checkout_table_row_number(n)
create_order_list_for_checkout()
orders = [x for x in orders if x is not None]

print(active_tables_nums)
print(index_nums)
print(orders)
