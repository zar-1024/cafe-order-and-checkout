from openpyxl import load_workbook

workbook = load_workbook("test_1.xlsx")
active_list = []

def add_input():
    # input the table number
    table_num = int(input("Input the table number: "))
    active_list.append(table_num)

    # input the order as string then add it to the order list
    order = input("add the order: ")
    active_list.append(order)

    # check if there are other orders for the same table
    add_order = input("do u want to add another order? y/n: ")

    # while loop for adding other orders to the same table
    while add_order == "y":
        order = input("add the order: ")
        active_list.append(order)
        add_order = input("do u want to add another order? y/n: ")



#function that adds items to the excel file
def active_tabs_input():
    add_input()
    sheet_1 = workbook['active_tabs']
    #max rows to decide the starting row as max_row+1
    num_rows = sheet_1.max_row
    #i = index so the items from the list take the same index on the data sheet
    for i, item in enumerate(active_list, start=0):
        sheet_1.cell(row=num_rows + 1 , column=i + 1).value = item
    workbook.save("test_1.xlsx")
    #clearing the list so we can start a new cycle
    active_list.clear()

def choose_task():
    choose_task = input("do u want to register an order or checkout? o/c ")
    match choose_task:
        case "o":
            active_tabs_input()
        case "c":
            print("sorry this function hasn't been activated yet")

#make the program work forever
while True:
    choose_task()